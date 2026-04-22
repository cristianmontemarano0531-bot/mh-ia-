import openpyxl
import json
import re
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(script_dir, "../base de datos productos completa.xlsx")

wb = openpyxl.load_workbook(excel_path, data_only=True)

def limpiar(val):
    if val is None:
        return ""
    s = str(val)
    reemplazos = {
        "\u00f1": "n", "\u00d1": "N",
        "\u00e9": "e", "\u00e1": "a", "\u00ed": "i", "\u00f3": "o", "\u00fa": "u",
        "\u00c9": "E", "\u00c1": "A", "\u00cd": "I", "\u00d3": "O", "\u00da": "U",
        "\u00fc": "u", "\u00e0": "a",
    }
    for bad, good in reemplazos.items():
        s = s.replace(bad, good)
    return s.strip()

def detectar_familia(codigo):
    c = codigo.upper()
    if c.endswith("COLOR"):
        return c[:-5]
    if c.endswith("CB"):
        return c[:-1]
    if c.endswith("B") and len(c) > 2:
        return c[:-1]
    return c

STOPWORDS = {"de", "con", "en", "y", "a", "el", "la", "los", "las", "un", "una", "para", "por", "cm", "al", "del"}

def extraer_keywords(texto):
    if not texto:
        return []
    kws = []
    for w in texto.lower().split():
        w = w.strip(".,;!?():")
        if w and len(w) > 1 and w not in STOPWORDS:
            kws.append(w)
    return kws

# ─── LEER SHEET1: 94 productos con enriquecimiento ───────────────────────────
# Columnas: Cod Producto | Producto | Rubro | Sub Rubro | DESCRIPCION LARGA |
#           PRODUCTOS RELACIONADO | FRASE GANADORA | colores | motor de busqueda |
#           tipo de instalacion | ideal para | descripcion corta | tags
ws1 = wb['Sheet1']
rows1 = list(ws1.iter_rows(min_row=2, values_only=True))

productos = {}
for r in rows1:
    codigo = str(r[0] or "").strip().upper()
    if not codigo:
        continue

    colores_raw = limpiar(r[7])
    relacionados_raw = limpiar(r[5])
    tags_raw = limpiar(r[12])
    motor_busqueda = limpiar(r[8])
    frase = limpiar(r[6])

    colores = [c.strip().lower() for c in re.split(r"[/,]", colores_raw) if c.strip()] if colores_raw else []

    relacionados = []
    for rel in relacionados_raw.split(","):
        rel = rel.strip().rstrip(".")
        if rel and not rel.upper().startswith("OPCIONAL") and not rel.upper().startswith("COMPATIBLE"):
            relacionados.append(rel.upper())

    tags = [t.strip().lower() for t in tags_raw.split(",") if t.strip()] if tags_raw else []

    motor_keywords = list(set(extraer_keywords(motor_busqueda) + extraer_keywords(frase)))

    familia = detectar_familia(codigo)

    productos[codigo] = {
        "codigo": codigo,
        "nombre": limpiar(r[1]),
        "rubro": limpiar(r[2]),
        "sub_rubro": limpiar(r[3]),
        "familia": familia,
        "colores_excel": colores,
        "motor_keywords": motor_keywords,
        "tags": tags,
        "relacionados": relacionados,
        "desc_larga": limpiar(r[4]),
        "tipo_instalacion": limpiar(r[9]),
        "ideal_para": limpiar(r[10]),
        "frase": frase,
        "fuente": "sheet1"
    }

# ─── LEER HOJA1: 389 productos completos de Dux ──────────────────────────────
# Columnas: CODIGO | PRODUCTO | RUBRO | SUB RUBRO | MARCA | UTILIZA VARIANTES | ...
RUBROS_RELEVANTES = {
    "MUEBLES", "BACHAS", "MESADAS", "ESPEJOS Y BOTIQUINES",
    "BLANCO LINEA", "U\xd1ERO", "PLACARD", "PLACARD 90",
    "UÑERO"
}

ws2 = wb['Hoja1']
rows2 = list(ws2.iter_rows(min_row=2, values_only=True))

nuevos_de_hoja1 = 0
for r in rows2:
    codigo = str(r[0] or "").strip().upper()
    if not codigo:
        continue
    rubro = str(r[2] or "").strip().upper()
    if rubro not in RUBROS_RELEVANTES and "U" + "\xd1" + "ERO" not in rubro and "U\xf1ERO" not in rubro:
        continue

    if codigo in productos:
        continue  # Ya enriquecido con Sheet1, no pisar

    nombre = limpiar(r[1])
    familia = detectar_familia(codigo)

    # Keywords basicos del nombre
    motor_keywords = extraer_keywords(nombre)

    productos[codigo] = {
        "codigo": codigo,
        "nombre": nombre,
        "rubro": rubro,
        "sub_rubro": limpiar(r[3]),
        "familia": familia,
        "colores_excel": [],
        "motor_keywords": motor_keywords,
        "tags": [],
        "relacionados": [],
        "desc_larga": "",
        "tipo_instalacion": "",
        "ideal_para": "",
        "frase": "",
        "fuente": "hoja1"
    }
    nuevos_de_hoja1 += 1

# ─── CONSTRUIR FAMILIAS ───────────────────────────────────────────────────────
familias = {}
for cod, data in productos.items():
    fam = data["familia"]
    if fam not in familias:
        familias[fam] = []
    familias[fam].append({
        "codigo": cod,
        "colores": data["colores_excel"],
        "es_blanco": "blanco" in data["colores_excel"],
        "es_color": any(c != "blanco" for c in data["colores_excel"])
    })

for cod, data in productos.items():
    data["variantes_familia"] = familias.get(data["familia"], [])

# ─── GUARDAR ─────────────────────────────────────────────────────────────────
output = {"productos": productos, "familias": familias}
out_path = os.path.join(script_dir, "base-mh.json")
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

sheet1_count = sum(1 for p in productos.values() if p["fuente"] == "sheet1")
hoja1_count = sum(1 for p in productos.values() if p["fuente"] == "hoja1")
familias_multi = [(f, vs) for f, vs in familias.items() if len(vs) > 1]

print(f"OK - {len(productos)} productos totales")
print(f"  Con enriquecimiento completo (Sheet1): {sheet1_count}")
print(f"  Con keywords basicos (Hoja1):          {hoja1_count}")
print(f"  Familias con variantes de color:       {len(familias_multi)}")
for fam, variantes in familias_multi:
    print(f"    {fam}: {[v['codigo'] for v in variantes]}")
